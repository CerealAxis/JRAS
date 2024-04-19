# coding='utf-8'
import argparse
import os
import requests
import json
import time
import openpyxl
from openpyxl import Workbook


def start(page, product_id, score):
    """
    发起请求并获取指定商品评分、页码的评论数据

    参数:
    - page (int): 页码
    - product_id (int): 商品ID
    - score (int): 评论类型，4为全部评论

    返回:
    - dict: 解析后的JSON数据
    """
    url = f'https://club.jd.com/comment/productPageComments.action?&productId={product_id}&score={score}&sortType=5&page={page}&pageSize=10&isShadowSku=0&fold=1'

    headers = {
        "User-Agent": "Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Mobile Safari/537.36"
    }
    time.sleep(2)
    response = requests.get(url=url, headers=headers)
    data = json.loads(response.text)
    return data


def parse(data):
    """
    解析评论数据，生成包含用户昵称、评论ID、评论内容、创建时间的元组序列

    参数:
    - data (dict): JSON格式的评论数据

    返回:
    - generator: 包含用户昵称、评论ID、评论内容、创建时间的元组序列
    """
    items = data['comments']
    for i in items:
        yield (
            i['nickname'],
            i['id'],
            i['content'],
            i['creationTime']
        )


def excel(items, product_id):
    """
    将评论数据写入Excel文件（如不存在则创建）

    参数:
    - items (generator): 包含用户昵称、评论ID、评论内容、创建时间的元组序列
    - product_id (int): 商品ID
    """
    if not os.path.exists('output'):
        os.makedirs('output')
    new_table = f'output/original_data.xlsx'
    wb = Workbook()
    ws = wb.active

    # 设置表头
    head_data = ['nickname', 'id', '内容', '时间']
    for i in range(0, 4):
        ws.cell(row=1, column=i + 1).value = head_data[i]

    index = 2

    # 写入评论数据
    for data in items:
        for i in range(0, 4):
            print(data[i])
            ws.cell(row=index, column=i + 1).value = data[i]
        print('______________________')
        index += 1

    wb.save(new_table)


def another(items, j, product_id):
    """
    将评论数据追加到已存在的Excel文件中

    参数:
    - items (generator): 包含用户昵称、评论ID、评论内容、创建时间的元组序列
    - j (int): 当前页码
    - product_id (int): 商品ID
    """
    if not os.path.exists('output'):
        os.makedirs('output')
    new_table = f'output/original_data.xlsx'
    index = (j - 1) * 10 + 2  # 计算起始行索引

    data = openpyxl.load_workbook(new_table)
    ws = data.active

    # 追加评论数据
    for test in items:
        for i in range(0, 4):
            print(test[i])
            ws.cell(row=index, column=i + 1).value = test[i]
        print('_______________________')
        index += 1

    data.save(new_table)


def main():
    """
    主程序：解析命令行参数，获取商品评论数据，并将数据写入Excel文件
    """
    parser = argparse.ArgumentParser(description="京东商品评论爬虫")
    parser.add_argument('-p', '--product-id', required=True, type=int,
                        help="输入商品ID")

    args = parser.parse_args()
    product_id = args.product_id
    score = 4
    page_amount = 20

    j = 1
    judge = True

    for i in range(0, page_amount):
        time.sleep(1.5)
        first = start(j, product_id, score)
        test = parse(first)

        if judge:
            excel(test, product_id)
            judge = False
        else:
            another(test, j, product_id)
        print(f'第{j}页抓取完毕')
        j += 1


if __name__ == '__main__':
    main()