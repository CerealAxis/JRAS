# coding='utf-8'
import openpyxl
import unicodedata
from collections import OrderedDict


def remove_duplicates_and_emojis(original_file, new_file):
    print("开始处理原始文件: ", original_file)

    wb = openpyxl.load_workbook(original_file)
    ws = wb.active

    # 读取表头
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active

    # 将表头复制到新工作簿中
    for i, header in enumerate(headers, start=1):
        ws_new.cell(row=1, column=i).value = header

    # 使用有序字典存储去重后的评论及其对应列数据
    comments_dict = OrderedDict()

    # 遍历原始数据，去除表情符号并去重
    for row in range(2, ws.max_row + 1):
        content = ws.cell(row=row, column=3).value
        if not content:
            continue

        content_no_emoji = remove_emojis(content)
        if content_no_emoji not in comments_dict:
            comments_dict[content_no_emoji] = [ws.cell(row=row, column=i).value for i in range(1, ws.max_column + 1)]

    # 将去重后的内容写入新工作簿
    for row_index, content in enumerate(comments_dict.values(), start=2):
        for col_index, cell_value in enumerate(content):
            if col_index == 2:  # 对第3列（即评论列）单独去除表情符号
                cell_value = remove_emojis(cell_value)
            ws_new.cell(row=row_index, column=col_index + 1).value = cell_value

    wb_new.save(new_file)

    print("数据清理完毕，表情已删除。结果保存至文件: ", new_file)


def remove_emojis(text):
    """
    移除文本中的表情符号

    参数:
    - text (str): 输入文本

    返回:
    - str: 去除表情符号后的文本
    """
    clean_text = ""
    for char in text:
        if unicodedata.category(char) != 'So':
            clean_text += char
    return clean_text


remove_duplicates_and_emojis('output/original_data.xlsx', 'output/cleaned_data.xlsx')
