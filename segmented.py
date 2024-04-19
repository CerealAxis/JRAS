# coding='utf-8'
import jieba
import jieba.analyse
import numpy as np
from PIL import Image
from openpyxl import load_workbook
from wordcloud import WordCloud
import matplotlib.pyplot as plt


def chinese_word_segmentation(file_path):
    """
    分词并生成词云

    参数:
    - file_path (str): Excel文件路径

    步骤：
    1. 从Excel文件中提取评论内容进行分词
    2. 过滤停用词并保存分词结果到文本文件
    3. 统计分词结果的词频
    4. 使用背景图片和词频数据生成词云
    5. 保存词云图片
    """
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 设置停用词
    jieba.analyse.set_stop_words("data/stopwords.txt")
    stop_words = set()
    with open('data/stopwords.txt', 'r', encoding='utf-8') as f:
        for word in f:
            stop_words.add(word.strip())

    # 从Excel文件中提取评论内容并分词
    segmented_words = []
    for row_index in range(2, sheet.max_row + 1):
        content = sheet.cell(row=row_index, column=3).value
        if content is not None:
            words = jieba.cut(content)
            words_filtered = ' '.join(word for word in words if word not in stop_words)
            segmented_words.append(words_filtered)

    # 保存分词结果到文本文件
    with open(f'output/segmented_words.txt', 'w', encoding='utf-8') as file:
        for words in segmented_words:
            file.write(words + '\n')
    print("Chinese word segmentation is completed and saved in 'segmented_words.txt'.")

    # 统计词频
    word_freq = {}
    for words in segmented_words:
        for word in words.split():
            if word not in stop_words:
                if word not in word_freq:
                    word_freq[word] = 0
                word_freq[word] += 1

    # 加载背景图片并生成词云
    img = Image.open("data/bg.png")
    mask = np.array(img)

    wordcloud = WordCloud(width=1000, height=1000, mask=mask, background_color='white', font_path='STKAITI.TTF',
                          stopwords=stop_words, random_state=50, max_words=40)
    wordcloud.generate_from_frequencies(word_freq)

    plt.imshow(wordcloud, interpolation='bilinear')
    plt.axis('off')
    wordcloud.to_file(f'output/wordcloud.png')


chinese_word_segmentation(f'output/cleaned_data.xlsx')
